"""生成选品对比表交付物：模板（xlsx）+ 示例（csv/md）。

用法：
    python tools/make_assets.py
"""
import csv
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from shopmonitor.collectors.mock import MockAdapter  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
ASSETS = ROOT / "assets"

_HEADERS = [
    "平台", "商品ID", "标题", "价格", "原价", "促销", "库存",
    "销量", "评分", "评论数", "店铺", "店铺评分", "排名", "链接", "利润估算", "备注",
]
_SAMPLE_NOTES = ["主推款，转化率高", "利润薄但走量", "新品，观察一周", "库存告急，慎压货", "竞品低价款", "适合做搭配购"]


def _rows(products):
    rows = []
    for i, p in enumerate(products, start=1):
        profit = round(p.price * 0.22, 2) if p.price else None
        rows.append(
            [
                p.platform, p.product_id, p.title, p.price, p.original_price, p.promo_text, p.stock_status,
                p.sales, p.rating, p.review_count, p.shop_name, p.shop_rating, p.rank, p.url, profit,
                _SAMPLE_NOTES[i % len(_SAMPLE_NOTES)],
            ]
        )
    return rows


def write_csv(rows):
    path = ASSETS / "选品对比表-示例.csv"
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(_HEADERS)
        w.writerows(rows)
    return path


def write_md(rows):
    path = ASSETS / "选品对比表-示例.md"
    lines = ["# 选品对比表（示例）", "", "> 由 ShopMonitor 生成。监控维度：价格/原价/促销/库存/销量/评分/评论数/店铺/排名。", ""]
    lines.append("| " + " | ".join(_HEADERS) + " |")
    lines.append("|" + "---|" * len(_HEADERS))
    for r in rows:
        cells = [str(x).replace("|", "\\|") if x is not None else "-" for x in r]
        lines.append("| " + " | ".join(cells) + " |")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


def write_xlsx(rows):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "选品对比表"
    ws.append(_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for r in rows:
        ws.append(r)
    widths = [10, 16, 46, 10, 10, 14, 8, 10, 8, 10, 14, 10, 8, 44, 12, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws2 = wb.create_sheet("使用说明")
    tips = [
        "选品对比表使用说明",
        "",
        "1. 数据来源：ShopMonitor 各平台榜单/搜索接口自动抓取，价格销量历史存 SQLite。",
        "2. 监控维度（电商运营核心需求）：",
        "   - 价格/原价/促销：发现降价、秒杀、领券（降价 24h 是流量重分配窗口）",
        "   - 库存：缺货/预售预警",
        "   - 销量与销量趋势：识别上升期，避开已爆品",
        "   - 评分/评论数：监控新增差评与评价波动",
        "   - 店铺/店铺评分：供应链与售后质量判断",
        "   - 排名：榜单/搜索位次变化",
        "3. 利润估算：示例列按售价*22% 粗估，请按实际进货价/运费/平台佣金修改公式。",
        "4. 选品筛选建议：",
        "   - 销量>1000 且价格带 50-300 元（大众快消）优先看",
        "   - 价格频繁变动 = 竞争激烈，慎入",
        "   - 评分<4.0 或差评集中在质量 = 供应链风险",
        "   - 店铺数少且搜索量大的关键词 = 蓝海机会",
        "5. 配套接口：GET /api/v1/rank/{platform}、GET /api/v1/report/compare、GET /api/v1/product/{platform}/{id}/change",
    ]
    for t in tips:
        ws2.append([t])
    ws2.column_dimensions["A"].width = 90

    path = ASSETS / "选品对比表-模板.xlsx"
    wb.save(path)
    return path


def main() -> None:
    ASSETS.mkdir(exist_ok=True)
    products = MockAdapter().fetch_rank(category="数码", limit=12)
    rows = _rows(products)
    print("CSV :", write_csv(rows))
    print("MD  :", write_md(rows))
    print("XLSX:", write_xlsx(rows))


if __name__ == "__main__":
    main()