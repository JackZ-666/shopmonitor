"""选品对比表生成：CSV / Markdown / Excel。"""
import csv
import io
from typing import List

from .cache import get_product
from .collectors.base import CollectorError
from .collectors.registry import get_adapter
from .models import CompareRow

# 按主流电商监控需求设计的对比表字段
_HEADERS = [
    "平台", "商品ID", "标题", "价格", "原价", "促销", "库存",
    "销量", "评分", "评论数", "店铺", "店铺评分", "排名", "链接", "预估毛利", "毛利率",
]
_HEADERS_EN = [
    "Platform", "Product ID", "Title", "Price", "Original", "Promo", "Stock",
    "Sales", "Rating", "Reviews", "Shop", "Shop Rating", "Rank", "URL", "Est. Profit", "Margin",
]


def _headers(lang: str) -> list:
    return _HEADERS_EN if lang == "en" else _HEADERS


def build_compare_rows(
    platform: str,
    product_ids: List[str],
    profit_cost_rate: float = 0.4,
    profit_shipping: float = 0.0,
    profit_acos: float = 0.0,
) -> List[CompareRow]:
    from .insights import estimate_item_profit

    rows: List[CompareRow] = []
    adapter = get_adapter(platform)
    for pid in product_ids:
        cached = get_product(platform, pid)
        if cached:
            r = _to_row(platform, pid, cached)
            r.estimated_profit, r.estimated_margin = estimate_item_profit(
                r.price, platform, profit_cost_rate, profit_shipping, profit_acos)
            rows.append(r)
            continue
        try:
            p = adapter.fetch_product(pid)
            est, margin = estimate_item_profit(
                p.price, platform, profit_cost_rate, profit_shipping, profit_acos)
            rows.append(
                CompareRow(
                    platform=p.platform,
                    product_id=p.product_id,
                    title=p.title,
                    price=p.price,
                    sales=p.sales,
                    shop_name=p.shop_name,
                    rating=p.rating,
                    review_count=p.review_count,
                    stock_status=p.stock_status,
                    promo_text=p.promo_text,
                    rank=p.rank,
                    url=p.url,
                    estimated_profit=est,
                    estimated_margin=margin,
                    crawled_at=p.crawled_at,
                )
            )
        except CollectorError:
            rows.append(CompareRow(platform=platform, product_id=pid, title="（未抓取到，见历史接口）"))
    return rows


def _to_row(platform: str, pid: str, cached: dict) -> CompareRow:
    return CompareRow(
        platform=platform,
        product_id=pid,
        title=cached.get("title", ""),
        price=cached.get("price"),
        sales=cached.get("sales"),
        shop_name=cached.get("shop_name"),
        rating=cached.get("rating"),
        review_count=cached.get("review_count"),
        stock_status=cached.get("stock_status"),
        promo_text=cached.get("promo_text"),
        rank=cached.get("rank"),
        url=cached.get("url"),
        crawled_at=cached.get("crawled_at", ""),
    )


def _cells(r: CompareRow) -> list:
    price = f"{r.price:.2f}" if r.price is not None else "-"
    orig = f"{r.original_price:.2f}" if getattr(r, "original_price", None) else "-"
    return [
        r.platform,
        r.product_id,
        r.title,
        price,
        orig,
        r.promo_text or "-",
        r.stock_status or "-",
        r.sales if r.sales is not None else "-",
        r.rating if r.rating is not None else "-",
        r.review_count if r.review_count is not None else "-",
        r.shop_name or "-",
        r.shop_rating if getattr(r, "shop_rating", None) else "-",
        r.rank if r.rank is not None else "-",
        r.url or "-",
        f"{r.estimated_profit:.2f}" if getattr(r, "estimated_profit", None) is not None else "-",
        f"{r.estimated_margin}%" if getattr(r, "estimated_margin", None) is not None else "-",
    ]


def to_markdown(rows: List[CompareRow], lang: str = "zh") -> str:
    hd = _headers(lang)
    lines = ["| " + " | ".join(hd) + " |", "|" + "---|" * len(hd)]
    for r in rows:
        cells = [str(x).replace("|", "\\|") for x in _cells(r)]
        lines.append("| " + " | ".join(cells) + " |")
    return "\n".join(lines) + "\n"


def to_csv(rows: List[CompareRow], lang: str = "zh") -> str:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(_headers(lang))
    for r in rows:
        writer.writerow(_cells(r))
    return buf.getvalue()


# ---------------- 选品毛利表导出（潜力筛选结果） ----------------
_PROFIT_HEADERS = ["排名", "商品ID", "标题", "价格", "销量", "评分", "潜力分", "预估毛利", "毛利率", "链接"]
_PROFIT_HEADERS_EN = ["Rank", "Product ID", "Title", "Price", "Sales", "Rating", "Score", "Est. Profit", "Margin", "URL"]


def _profit_cells(it: dict, lang: str) -> list:
    price = f"{it['price']:.2f}" if it.get("price") is not None else "-"
    profit = it.get("estimated_profit")
    margin = it.get("estimated_margin")
    return [
        it.get("rank") if it.get("rank") is not None else "-",
        it.get("product_id") or "-",
        (it.get("title") or "-").replace("|", "\\|"),
        price,
        it.get("sales") if it.get("sales") is not None else "-",
        it.get("rating") if it.get("rating") is not None else "-",
        it.get("score") if it.get("score") is not None else "-",
        f"{profit:.2f}" if profit is not None else "-",
        f"{margin}%" if margin is not None else "-",
        it.get("url") or "-",
    ]


def export_profit_table(items: list, fmt: str = "md", path: str = "", lang: str = "zh") -> str:
    """把筛选结果导出为选品毛利表：fmt=md|csv|xlsx。xlsx 需传 path。"""
    hd = _PROFIT_HEADERS_EN if lang == "en" else _PROFIT_HEADERS
    if fmt == "pdf":
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle

        style = _pdf_style()
        data = [[Paragraph(h, style) for h in hd]]
        for it in items:
            data.append([Paragraph(str(c), style) for c in _profit_cells(it, lang)])
        doc = SimpleDocTemplate(path, pagesize=landscape(A4),
                                leftMargin=10 * mm, rightMargin=10 * mm,
                                topMargin=10 * mm, bottomMargin=10 * mm, title="选品毛利表")
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#cccccc")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        doc.build([table])
        return path
    if fmt == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(hd)
        for it in items:
            writer.writerow(_profit_cells(it, lang))
        return buf.getvalue()
    if fmt == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "选品毛利表"
        ws.append(hd)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
        for it in items:
            ws.append(_profit_cells(it, lang))
        for i, _ in enumerate(hd, start=1):
            ws.column_dimensions[get_column_letter(i)].width = 14
        ws.column_dimensions["C"].width = 50
        ws.column_dimensions["J"].width = 44
        wb.save(path)
        return path
    lines = ["# 选品毛利表", "",
             "| " + " | ".join(hd) + " |", "|" + "---|" * len(hd)]
    for it in items:
        cells = [str(x).replace("|", "\\|") for x in _profit_cells(it, lang)]
        lines.append("| " + " | ".join(cells) + " |")
    return "\n".join(lines) + "\n"


def _pdf_style():
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont

    try:
        pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    except Exception:  # noqa: BLE001
        pass
    return ParagraphStyle(
        "zh", fontName="STSong-Light", fontSize=8, leading=11,
        textColor=colors.HexColor("#222222"),
    )


def to_pdf(rows: List[CompareRow], path: str, lang: str = "zh") -> str:
    """选品对比表 -> PDF（横向，中文用 STSong-Light 内嵌字体）。"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    hd = _headers(lang)
    style = _pdf_style()
    data = [[Paragraph(h, style) for h in hd]]
    for r in rows:
        data.append([Paragraph(str(c), style) for c in _cells(r)])
    doc = SimpleDocTemplate(
        path, pagesize=landscape(A4),
        leftMargin=10 * mm, rightMargin=10 * mm, topMargin=10 * mm, bottomMargin=10 * mm,
        title="选品对比表",
    )
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#cccccc")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f6f9")]),
    ]))
    doc.build([table])
    return path


def md_to_pdf(md: str, path: str, title: str = "ShopMonitor 报告") -> str:
    """Markdown 文本 -> PDF（简单渲染：标题/列表/段落/表格行）。"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

    try:
        pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    except Exception:  # noqa: BLE001
        pass
    title_style = ParagraphStyle("t", fontName="STSong-Light", fontSize=15, leading=20,
                                 textColor=colors.HexColor("#123c6e"), spaceAfter=8)
    h_style = ParagraphStyle("h", fontName="STSong-Light", fontSize=11, leading=15,
                             textColor=colors.HexColor("#1b5bbf"), spaceBefore=8, spaceAfter=4)
    b_style = ParagraphStyle("b", fontName="STSong-Light", fontSize=9, leading=13)
    story = [Paragraph(title, title_style)]
    for line in md.splitlines():
        s = line.strip()
        if not s:
            continue
        if s.startswith("# "):
            story.append(Paragraph(s[2:], title_style))
        elif s.startswith("## "):
            story.append(Paragraph(s[3:], h_style))
        elif s.startswith("### "):
            story.append(Paragraph(s[4:], h_style))
        elif s.startswith("- ") or s.startswith("* "):
            story.append(Paragraph("• " + s[2:], b_style))
        elif s.startswith("|"):
            story.append(Paragraph(s.replace("|", "  "), b_style))
        else:
            story.append(Paragraph(s, b_style))
        story.append(Spacer(1, 2))
    doc = SimpleDocTemplate(path, pagesize=A4, leftMargin=14 * mm, rightMargin=14 * mm,
                            topMargin=14 * mm, bottomMargin=14 * mm, title=title)
    doc.build(story)
    return path


def to_excel(rows: List[CompareRow], path: str, lang: str = "zh") -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    hd = _headers(lang)
    wb = Workbook()
    ws = wb.active
    ws.title = "Compare" if lang == "en" else "选品对比表"
    ws.append(hd)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
    for r in rows:
        ws.append(_cells(r))
    for i, _ in enumerate(hd, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["N"].width = 50
    wb.save(path)
    return path